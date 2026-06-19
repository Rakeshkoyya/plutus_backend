"""Generic xlsx ingestion with AI-assisted column mapping via OpenRouter."""
import json
import re
from typing import Any

import httpx
from openpyxl import load_workbook

from app.config import settings

# Target schemas the importer can populate
TARGET_FIELDS = {
    "students_fees": [
        # identity
        "name", "admission_number", "roll_number", "class_name", "section", "category",
        # contacts
        "father_name", "father_phone", "mother_name", "mother_phone", "phone",
        # fees
        "total_fee", "fee_after_discount", "discount", "opening_dues",
        # collections — either a single paid amount, or per-quarter collected amounts
        "paid_amount", "q1_paid", "q2_paid", "q3_paid", "q4_paid",
        "receipt_number",
    ],
    "fee_structures": [
        "class_name", "academic_year", "total_amount", "num_installments",
    ],
}

# Quarter "paid/received" columns the importer recognises as quarterly collections.
QUARTER_FIELDS = ["q1_paid", "q2_paid", "q3_paid", "q4_paid"]
QUARTER_LABELS = ["1st Quarter", "2nd Quarter", "3rd Quarter", "4th Quarter"]

FIELD_HINTS = {
    "name": ["student name", "name", "student", "full name"],
    "admission_number": ["sr no", "sr. no", "adm", "admission", "adm. no", "scholar"],
    "roll_number": ["roll number", "roll no", "roll"],
    "class_name": ["class", "grade", "standard", "std", "class name"],
    "section": ["section", "sec", "div", "division"],
    "category": ["category", "student category", "type", "status"],
    "father_name": ["father's name", "father name", "father", "guardian"],
    "father_phone": ["father's mobile", "father mobile", "father's phone", "father phone"],
    "mother_name": ["mother's name", "mother name", "mother"],
    "mother_phone": ["mother's mobile", "mother mobile", "mother's phone", "mother phone"],
    "phone": ["contact number", "contact", "mobile", "phone", "mobile no"],
    "total_fee": ["total fee", "total", "annual fee", "tuition", "total amount"],
    "fee_after_discount": ["fee after discount", "net fee", "payable", "after discount"],
    "discount": ["discount", "concession", "waiver", "scholarship"],
    "opening_dues": ["last yr due", "last year due", "previous due", "arrears", "old due", "opening"],
    "paid_amount": ["paid", "amount paid", "received", "collected", "total paid"],
    "q1_paid": ["1st qtr paid", "1st qtr received", "q1 paid", "1st quarter", "qtr 1"],
    "q2_paid": ["2nd qtr paid", "2nd qtr received", "q2 paid", "2nd quarter", "qtr 2"],
    "q3_paid": ["3rd qtr paid", "3rd qtr received", "q3 paid", "3rd quarter", "qtr 3"],
    "q4_paid": ["4th qtr paid", "4th qtr received", "q4 paid", "4th quarter", "qtr 4"],
    "receipt_number": ["payment details", "receipt", "receipt no", "receipt number", "fr no", "voucher"],
    "academic_year": ["academic year", "year", "session", "ay"],
    "total_amount": ["total fee", "total amount", "total", "fee", "amount"],
    "num_installments": ["installments", "installment", "no of installments", "terms"],
}


# ---------------------------------------------------------------------------
# Academic-year detection
# ---------------------------------------------------------------------------
# The school's files usually carry the year in the file name (shana_fee_2024-25.xlsx)
# or a sheet tab ("Class 5 2024-25"). We try those first; if nothing is found the
# router raises a clarification question for the user to answer.

# Explicit range: 2023-24, 2023-2024, 2023/24, 2023_2024 (en/em dashes too).
_AY_RANGE_RE = re.compile(r"(20\d{2})\s*[-_/–—]\s*(20\d{2}|\d{2})")
# A lone 4-digit calendar year — only trusted in names / year-headed columns.
_AY_SINGLE_RE = re.compile(r"(?<!\d)(20\d{2})(?!\d)")

YEAR_HEADER_HINTS = ["academic year", "acad year", "session", "ay", "year"]


def _format_ay(start: int) -> str:
    return f"{start}-{(start + 1) % 100:02d}"


def _ay_from_range(start: int, end: int) -> str:
    if end < 100:  # two-digit end like "24" -> 2024
        end = (start // 100) * 100 + end
        if end <= start:
            end += 100
    # Academic years span exactly one calendar year; normalise to start-(start+1).
    return _format_ay(start)


def detect_ay_in_text(text: Any, allow_single: bool = True) -> str | None:
    """Pull an academic year like '2023-24' out of a file name or sheet title."""
    if text is None:
        return None
    s = str(text)
    m = _AY_RANGE_RE.search(s)
    if m:
        return _ay_from_range(int(m.group(1)), int(m.group(2)))
    if allow_single:
        m = _AY_SINGLE_RE.search(s)
        if m and 2000 <= int(m.group(1)) <= 2099:
            return _format_ay(int(m.group(1)))
    return None


def detect_ay_in_rows(columns: list[str], sample_rows: list[dict]) -> str | None:
    """Use a column that holds a single, uniform academic-year value."""
    if not sample_rows:
        return None
    for col in columns:
        if not any(h in col.lower() for h in YEAR_HEADER_HINTS):
            continue
        values = [r.get(col) for r in sample_rows if r.get(col) not in (None, "")]
        found = {detect_ay_in_text(v, allow_single=True) for v in values}
        found.discard(None)
        if len(found) == 1:
            return next(iter(found))
    return None


def detect_academic_year(
    sheet_name: str, columns: list[str], sample_rows: list[dict], filename: str | None
) -> tuple[str | None, str | None]:
    """Best-effort year for a sheet. Returns (year, source) where source is one of
    'sheet', 'column', 'filename', or (None, None) when undetermined."""
    ay = detect_ay_in_text(sheet_name)
    if ay:
        return ay, "sheet"
    ay = detect_ay_in_rows(columns, sample_rows)
    if ay:
        return ay, "column"
    ay = detect_ay_in_text(filename)
    if ay:
        return ay, "filename"
    return None, None


def build_year_options(default_year: str, detected: list[str]) -> list[str]:
    """A sorted set of plausible years to offer in the clarification UI: a window
    around the default year plus a window around anything we detected."""
    opts: set[str] = set()
    for y in [default_year, *detected]:
        try:
            start = int(str(y)[:4])
        except (ValueError, TypeError):
            continue
        for s in range(start - 2, start + 3):
            opts.add(_format_ay(s))
    return sorted(opts)


def read_workbook(path: str, sample_size: int = 6) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # find first non-empty row as header
        header_idx = 0
        for idx, r in enumerate(rows):
            if any(c is not None and str(c).strip() for c in r):
                header_idx = idx
                break
        header = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(rows[header_idx])]
        data_rows = rows[header_idx + 1 :]
        records = []
        for r in data_rows:
            if not any(c is not None and str(c).strip() for c in r):
                continue
            rec = {}
            for i, h in enumerate(header):
                rec[h] = r[i] if i < len(r) else None
            records.append(rec)
        sheets.append(
            {
                "sheet_name": ws.title,
                "columns": [h for h in header if h],
                "rows": records,
            }
        )
    wb.close()
    return sheets


def _jsonsafe(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (int, float, str, bool)):
        return v
    return str(v)


def heuristic_mapping(columns: list[str]) -> dict[str, Any]:
    """Fallback when AI is unavailable: keyword-match columns to fields."""
    lc = {c.lower().strip(): c for c in columns}

    def find(field: str):
        for hint in FIELD_HINTS.get(field, []):
            for low, orig in lc.items():
                if hint == low or hint in low:
                    return orig
        return None

    students = {f: find(f) for f in TARGET_FIELDS["students_fees"]}
    students = {k: v for k, v in students.items() if v}
    # "Fee After Discount" contains the word "discount"; don't also claim it as a
    # separate discount column — the discount is derived from total − net instead.
    if students.get("discount") and students.get("discount") == students.get("fee_after_discount"):
        students.pop("discount", None)
    structures = {f: find(f) for f in TARGET_FIELDS["fee_structures"]}
    structures = {k: v for k, v in structures.items() if v}

    # Decide entity: fee_structures needs class + a total but few/no per-student name column
    if "name" in students:
        return {"entity": "students_fees", "mapping": students, "confidence": "low",
                "notes": "Heuristic match (AI disabled)."}
    if structures.get("class_name"):
        return {"entity": "fee_structures", "mapping": structures, "confidence": "low",
                "notes": "Heuristic match (AI disabled)."}
    return {"entity": "unknown", "mapping": {}, "confidence": "low",
            "notes": "Could not confidently map columns; please map manually."}


async def ai_mapping(sheet_name: str, columns: list[str], sample_rows: list[dict]) -> dict[str, Any] | None:
    if not settings.ai_enabled:
        return None
    prompt = f"""You are a data-mapping assistant for a school fee management system.

Given a spreadsheet sheet, decide which target entity it represents and map its columns
to our target fields.

TARGET ENTITIES & FIELDS:
1. "students_fees" (one row per student with their fee): {TARGET_FIELDS['students_fees']}
   - name is REQUIRED for this entity.
2. "fee_structures" (per-class fee template): {TARGET_FIELDS['fee_structures']}

SHEET NAME: {sheet_name}
COLUMNS: {columns}
SAMPLE ROWS (up to 6): {json.dumps(sample_rows[:6], default=str)}

Return ONLY a JSON object (no markdown) of the form:
{{"entity": "students_fees" | "fee_structures" | "unknown",
  "mapping": {{ "<target_field>": "<exact source column name>" }},
  "confidence": "high" | "medium" | "low",
  "notes": "<short explanation>"}}
Only include target fields you are confident about. Use exact source column names from COLUMNS."""

    headers = {
        "Authorization": f"Bearer {settings.openrouter_api_key}",
        "Content-Type": "application/json",
        "HTTP-Referer": settings.frontend_url,
        "X-Title": "Plutus — School Fee Management",
    }
    payload = {
        "model": settings.openrouter_model,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0,
        "response_format": {"type": "json_object"},
    }
    try:
        async with httpx.AsyncClient(timeout=40) as client:
            resp = await client.post(
                f"{settings.openrouter_base_url}/chat/completions",
                headers=headers,
                json=payload,
            )
        if resp.status_code != 200:
            return None
        content = resp.json()["choices"][0]["message"]["content"]
        match = re.search(r"\{.*\}", content, re.DOTALL)
        if not match:
            return None
        parsed = json.loads(match.group(0))
        # keep only valid source columns
        ent = parsed.get("entity", "unknown")
        valid = set(TARGET_FIELDS.get(ent, []))
        mapping = {
            k: v for k, v in (parsed.get("mapping") or {}).items()
            if k in valid and v in columns
        }
        return {
            "entity": ent,
            "mapping": mapping,
            "confidence": parsed.get("confidence", "medium"),
            "notes": parsed.get("notes", ""),
        }
    except Exception:
        return None
